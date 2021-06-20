from tkinter import*  #Importamos libreria tkinter
from os import listdir, popen
import PIL
from PIL import Image
import sys
import PyPDF2
from PyPDF2 import PdfFileReader, PdfFileWriter
import openpyxl as op
import time





LibrosLeidosTotales= op.load_workbook("Libros leidos totales.xlsx")
HojaLibrosLeidos=LibrosLeidosTotales["Libros"]

EtiquetasTotales=[]

for n in range (0, HojaLibrosLeidos.max_column-9):
	
	EtiquetaActual=HojaLibrosLeidos.cell(1,10+n).value
	EtiquetasTotales.append(EtiquetaActual)

print(EtiquetasTotales)



class ScrollFrame (Frame):
	def __init__(self, parent, bg="red", width=200, *args, **kwargs):
		super().__init__(parent,bg=bg, width=width, *args, **kwargs)
		self.CanvasPropio=Canvas(self, bg=bg, width=width)
		self.ScrollbarPropia=Scrollbar(self, orient="vertical", command=self.CanvasPropio.yview)
		self.FrameInside=Frame(self.CanvasPropio, bg=bg)

		self.FrameInside.bind("<Configure>",lambda e: self.CanvasPropio.configure(scrollregion=self.CanvasPropio.bbox("all")))
		self.CanvasPropio.create_window((0, 0), window=self.FrameInside, anchor="nw")
		self.CanvasPropio.configure(yscrollcommand=self.ScrollbarPropia.set)

		self.pack(side= LEFT, fill="both", expand=True)
		self.CanvasPropio.pack(side= LEFT, fill="both", expand=True)
		self.ScrollbarPropia.pack(side= RIGHT, fill="y")



class frametoTitleText(Frame):
#Packear y poner la misma fuente
#name_txt tiene que ser la direccion y el nombre
	def __init__(self, container, name_txt, fuente, w_title=10, w_entry=10,bg="white", starting_text='Clica aqui para agregar sinopsis', *args, **kwargs):
	    super().__init__(container, bg=bg,*args, **kwargs)
	    self.pack(fill="x")
	    
	    self.title = Label(self,text=starting_text, font=fuente, bg=bg,width=w_title)
	    self.title.pack(side="top", fill="x",pady=5)
	    self.name_txt=name_txt

	    self.text = Text(self, height=50, width=w_entry)
	    self.text.insert(INSERT, starting_text)
	    
	    self.title.bind('<Button-1>',lambda event: self.start(event))
	    self.text.bind('<Escape>',lambda event: self.end(event))

	def start(self,event):
		self.title.pack_forget()
		self.text.pack(side="top",fill="both", pady=6)#,fill="x"
		return

	def end(self,event):
		texto_editado=self.text.get("1.0", "end")
		with open(self.name_txt+".txt","w", encoding="utf8") as g:
			g.write(texto_editado)
			g.close()

		self.text.pack_forget()
		self.title.config(text=texto_editado)
		self.title.pack(side="top",fill="x",pady=5)
		return



class Checkbox(Checkbutton):
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.variable =BooleanVar(self)
        self.configure(variable=self.variable)
    
    def checked(self):
        return self.variable.get()
    
    def check(self):
        self.variable.set(True)
    
    def uncheck(self):
        self.variable.set(False)


#-----------------------------------------------------------------------------------------------------------


def actualizar(criterio, etiqueta):

	if etiqueta in criterio:
		criterio.pop(criterio.index(etiqueta))
		if len(criterio)==0:
			criterio.append("Todos")
		
	else:
		criterio.append(etiqueta)
		if "Todos" in criterio:
			criterio.pop(criterio.index("Todos"))
		
	ordenar(criterio)

	return 

def LibroEnCriterio(nLibro, criterio):
	EtiquetasLibro=[]
	EtiquetasExcel=[]
	CriterioLibro=False

	for k in range(0, HojaLibrosLeidos.max_column-9):
		EtiquetasExcel.append(HojaLibrosLeidos.cell(nLibro,10+k).value)
		if (EtiquetasExcel[k]=="Si"):
			EtiquetasLibro.append(EtiquetasTotales[k])


	for etiquetaCriterio in range (0, len(criterio)):
			for etiquetaCadaLibro in range (0, len(EtiquetasLibro)):
				if (EtiquetasLibro[etiquetaCadaLibro]==criterio[etiquetaCriterio]):
					CriterioLibro=True
	return CriterioLibro


def LibroEnEtiqueta(nLibro, netiqueta):
	EtiquetasLibro=[]
	EtiquetasExcel=[]
	EtiquetaLibro=False

	for k in range(0, HojaLibrosLeidos.max_column-9):
		EtiquetasExcel.append(HojaLibrosLeidos.cell(nLibro,10+k).value)
		if (EtiquetasExcel[k]=="Si"):
			EtiquetasLibro.append(EtiquetasTotales[k])


	for etiquetaCadaLibro in range (0, len(EtiquetasLibro)):
		if (EtiquetasLibro[etiquetaCadaLibro]==HojaLibrosLeidos.cell(1,netiqueta).value):
			EtiquetaLibro=True

	return EtiquetaLibro


def AgregarEtiquetaLibro(nLibro, netiqueta):
	valorEtiqueta=HojaLibrosLeidos.cell(nLibro,netiqueta).value
	if(valorEtiqueta==None):
		valorEtiqueta="Si"
		HojaLibrosLeidos.cell(nLibro,netiqueta).value=valorEtiqueta

	else:
		valorEtiqueta=None
		HojaLibrosLeidos.cell(nLibro,netiqueta).value=valorEtiqueta
	LibrosLeidosTotales.save("Libros leidos totales.xlsx")

	return


def ordenar(criterio):

	for widget in frame_scroll_libros.FrameInside.winfo_children():
		widget.destroy()


	nLibro=2
	columna=0
	fila=0
	libros=[]
	numLibro=0
	j=0
	k=0
	print(criterio)

	while numLibro<HojaLibrosLeidos.max_row-1:
		

		titulo=HojaLibrosLeidos.cell(nLibro,1).value
		autor=HojaLibrosLeidos.cell(nLibro,2).value
		saga=HojaLibrosLeidos.cell(nLibro,3).value
		mes=HojaLibrosLeidos.cell(nLibro,4).value
		año=HojaLibrosLeidos.cell(nLibro,5).value
		sinopsis=HojaLibrosLeidos.cell(nLibro,6).value
		portadaPeque=HojaLibrosLeidos.cell(nLibro,7).value
		portadaGrande=HojaLibrosLeidos.cell(nLibro,8).value
		NumPag=HojaLibrosLeidos.cell(nLibro,9).value



		if (LibroEnCriterio(nLibro, criterio)==True):
		
			ImagenPortadaPeque=PhotoImage(file=portadaPeque)

			espacioColumna1=Label(frame_scroll_libros.FrameInside, text="  ", bg="#FEDEC8")
			espacioColumna1.grid(row=fila, column=columna)
					
			libros.append(Button(frame_scroll_libros.FrameInside, image=ImagenPortadaPeque, command=lambda numeroLibro=nLibro: verInfo(numeroLibro)))
			libros[j].image=ImagenPortadaPeque
			libros[j].grid(row=fila, column=columna+1)

			espacioColumna2=Label(frame_scroll_libros.FrameInside, text="        ", bg="#FEDEC8")
			espacioColumna2.grid(row=fila, column=columna+2)

			espacioFila2=Label(frame_scroll_libros.FrameInside, text="        ", bg="#FEDEC8")
			espacioFila2.grid(row=fila+1, column=columna+1)


			

			if(columna<10):
				columna=columna+2

			else:
				columna=0
				fila=fila+2

			j=j+1
		numLibro=numLibro+1
		nLibro=nLibro+1

def verInfo(nLibro):

	for widget in frame_scroll_datos.FrameInside.winfo_children():
		widget.destroy()


	titulo=HojaLibrosLeidos.cell(nLibro,1).value
	autor=HojaLibrosLeidos.cell(nLibro,2).value
	saga=HojaLibrosLeidos.cell(nLibro,3).value
	mes=HojaLibrosLeidos.cell(nLibro,4).value
	año=HojaLibrosLeidos.cell(nLibro,5).value
	sinopsis=HojaLibrosLeidos.cell(nLibro,6).value
	portadaPeque=HojaLibrosLeidos.cell(nLibro,7).value
	portadaGrande=HojaLibrosLeidos.cell(nLibro,8).value
	NumPag=HojaLibrosLeidos.cell(nLibro,9).value

	ImagenPortadaGrande=PhotoImage(file=portadaGrande)
			
	portadaLabel=Label(frame_scroll_datos.FrameInside, image=ImagenPortadaGrande)
	portadaLabel.image=ImagenPortadaGrande
	portadaLabel.pack(padx=35, pady=10)

	tituloLibro=Label(frame_scroll_datos.FrameInside, text=titulo)
	tituloLibro.config(font=("Please write me a song", 30), justify="center",bg="#FEDEC8", padx=5, pady=15)
	tituloLibro.pack(padx=35, pady=5)

	autorLibro=Label(frame_scroll_datos.FrameInside, text=autor)
	autorLibro.config(font=("Please write me a song", 20), justify="center",bg="#FEDEC8", padx=5, pady=5)
	autorLibro.pack(padx=35, pady=5)

	if (saga!=" "):
		sagaLibro=Label(frame_scroll_datos.FrameInside, text=saga)
		sagaLibro.config(font=("Please write me a song", 20), justify="center",bg="#FEDEC8", padx=5, pady=5)
		sagaLibro.pack(padx=35, pady=5)


	numeroPag=Label(frame_scroll_datos.FrameInside, text="Número de páginas: "+str(NumPag))
	numeroPag.config(font=("Please write me a song", 20), justify="center",bg="#FEDEC8", padx=5, pady=5)
	numeroPag.pack(padx=35, pady=5)

	sinopsisLabel=Label(frame_scroll_datos.FrameInside, text=sinopsis)
	sinopsisLabel.config(font=("Century Gothic", 12), justify="center",bg="#FEDEC8", pady=10)
	sinopsisLabel.pack(pady=10)

	for etiqueta in range(0,len(EtiquetasTotales)):
		netiqueta=10+etiqueta
		checkbox = Checkbox(frame_scroll_datos.FrameInside, text=EtiquetasTotales[etiqueta], command=lambda libro=nLibro, netiqueta=netiqueta: AgregarEtiquetaLibro(libro, netiqueta) )
		checkbox.pack(side="top")
		checkbox.configure( bg="#FEDEC8")
		if(LibroEnEtiqueta(nLibro, netiqueta)==True):
			checkbox.check()        


#-------------------------------------------------------------------------------------------------------------




#Creamos la raiz (ventana), le damos titulo, icono y le decimos que está permitido cambiar de tamaño

raiz=Tk()
raiz.title("Bookshelf")
raiz.configure(bg="#FEDEC8")
raiz.geometry('1000x480')
raiz.resizable(True, True)
raiz.state("zoomed")



#Le ponemos un titulo (texto), un cuadro de texto y un boton para buscar

cabecera=Frame(raiz,bg="#FEDEC8")
cabecera.pack(side= TOP, fill=X)

titulo=Label(cabecera, text="Bookshelf: tu estantería virtual.")
titulo.pack(side=LEFT, pady=20, padx=30)
titulo.config(font=("hello honey - Personal use", 36), bg="#FEDEC8")


botonBuscar=Button(cabecera,text="Buscar", command="search")
botonBuscar.pack(side=RIGHT,pady=10, padx=10)

buscarTexto=Entry(cabecera)
buscarTexto.pack(side=RIGHT)

etiquetasFrame=Frame(raiz)
etiquetasFrame.config(bg="#FEDEC8", width=300)
etiquetasFrame.pack(side=LEFT, fill="y", expand=False)
categoriasLabel=Label(etiquetasFrame, text="Categorias")
categoriasLabel.config(font=("Please write me a song", 20), bg="#FEDEC8")
categoriasLabel.pack()

frame_scroll_libros=ScrollFrame(raiz, bg="#FEDEC8", width=800)
frame_scroll_datos=ScrollFrame(raiz, bg="#FEDEC8", width=400)	


criterio=["Todos"]
i=0

for i in range (0, HojaLibrosLeidos.max_column-10):
	
	Etiqueta=HojaLibrosLeidos.cell(1,11+i).value #"Universidad"

	BotonEtiqueta=Button(etiquetasFrame, text=Etiqueta, command=lambda criterio=criterio , etiqueta=Etiqueta: actualizar(criterio, etiqueta))
	BotonEtiqueta.pack()

ordenar(criterio)

raiz.mainloop()
